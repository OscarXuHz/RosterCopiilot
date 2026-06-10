"""Excel export helpers for the mock MVP.

The real product will need to preserve the NGO's existing workbook styling.
For the mock MVP we export an Excel-like matrix plus review sheets so the demo
can prove the end-to-end flow.
"""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import AuditItem, ExportRequest, ScheduleEntry, ScheduleResult
from .mock_scheduler import get_current_result

PERIOD_LABEL = {"AM": "上午", "PM": "下午"}


def _entry_label(entry: ScheduleEntry) -> str:
    parts = [entry.service_code.value]
    if entry.elder_name:
        parts.append(f":{entry.elder_name}")
    if entry.destination:
        parts.append(f"->{entry.destination}")
    if entry.status != "scheduled":
        parts.append(f"[{entry.status}]")
    if entry.notes:
        parts.append(f"\n{entry.notes}")
    return "".join(parts)


def _style_header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2563EB")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_subheader(cell) -> None:
    cell.font = Font(bold=True, color="111827")
    cell.fill = PatternFill("solid", fgColor="DBEAFE")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_border(ws) -> None:
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _add_schedule_sheet(wb: Workbook, result: ScheduleResult) -> None:
    ws = wb.active
    ws.title = "總排班"
    ws.freeze_panes = "C3"

    ws.cell(1, 1, "RosterCopiilot 排班建議")
    ws.cell(1, 1).font = Font(bold=True, size=16)
    ws.cell(1, 3, f"Week start: {result.week_start.isoformat()}")

    worker_names = sorted({e.worker_name for e in result.entries if e.worker_name and e.worker_name != "待分配"})
    headers = ["日期", "時段", *worker_names]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(2, col, header)
        _style_header(cell)

    by_slot: dict[tuple[str, str, str], list[ScheduleEntry]] = defaultdict(list)
    dates = sorted({e.schedule_date for e in result.entries})
    for entry in result.entries:
        if entry.worker_name == "待分配":
            continue
        by_slot[(entry.schedule_date.isoformat(), entry.period, entry.worker_name)].append(entry)

    row_idx = 3
    for service_date in dates:
        for period in ("AM", "PM"):
            ws.cell(row_idx, 1, service_date.isoformat())
            ws.cell(row_idx, 2, PERIOD_LABEL[period])
            for col_idx, worker in enumerate(worker_names, start=3):
                labels = [_entry_label(e) for e in by_slot.get((service_date.isoformat(), period, worker), [])]
                cell = ws.cell(row_idx, col_idx, "\n---\n".join(labels))
                if any("[needs_review]" in label or "[affected]" in label for label in labels):
                    cell.fill = PatternFill("solid", fgColor="FEF3C7")
                if any("[cancelled]" in label or "[unassigned]" in label for label in labels):
                    cell.fill = PatternFill("solid", fgColor="FECACA")
            row_idx += 1

    widths = [14, 10, *([24] * len(worker_names))]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in range(3, row_idx):
        ws.row_dimensions[row].height = 58
    _apply_border(ws)


def _add_audit_sheet(wb: Workbook, audit_items: list[AuditItem]) -> None:
    ws = wb.create_sheet("人工審核")
    headers = ["狀態", "類型", "嚴重度", "原因", "原安排", "建議安排", "人工備註"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        _style_header(cell)

    for row_idx, item in enumerate(audit_items, start=2):
        ws.cell(row_idx, 1, item.status.value)
        ws.cell(row_idx, 2, item.kind)
        ws.cell(row_idx, 3, item.severity)
        ws.cell(row_idx, 4, item.reason)
        ws.cell(row_idx, 5, _entry_label(item.original_entry) if item.original_entry else "")
        ws.cell(row_idx, 6, _entry_label(item.suggested_entry) if item.suggested_entry else "")
        ws.cell(row_idx, 7, item.human_note or "")

    for col_idx, width in enumerate([12, 22, 10, 48, 34, 34, 28], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    _apply_border(ws)


def _add_unassigned_sheet(wb: Workbook, unassigned: list[ScheduleEntry]) -> None:
    ws = wb.create_sheet("未分配")
    headers = ["日期", "時段", "服務", "長者", "區域", "目的地", "原因"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        _style_header(cell)

    for row_idx, entry in enumerate(unassigned, start=2):
        ws.cell(row_idx, 1, entry.schedule_date.isoformat())
        ws.cell(row_idx, 2, PERIOD_LABEL[entry.period])
        ws.cell(row_idx, 3, entry.service_code.value)
        ws.cell(row_idx, 4, entry.elder_name or "")
        ws.cell(row_idx, 5, entry.district or "")
        ws.cell(row_idx, 6, entry.destination or "")
        ws.cell(row_idx, 7, entry.notes or "")

    for col_idx, width in enumerate([14, 10, 14, 18, 14, 24, 44], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    _apply_border(ws)


def build_workbook(result: ScheduleResult, include_audit: bool = True) -> Workbook:
    wb = Workbook()
    _add_schedule_sheet(wb, result)
    if include_audit:
        _add_audit_sheet(wb, result.audit_items)
        _add_unassigned_sheet(wb, result.unassigned)
    for ws in wb.worksheets:
        for cell in ws[1]:
            if cell.value:
                _style_subheader(cell)
    return wb


def save_schedule_export(
    request: ExportRequest | None = None,
    output_dir: Path | None = None,
) -> Path:
    result = request.result if request and request.result else get_current_result()
    include_audit = request.include_audit if request else True
    output_dir = output_dir or Path(__file__).resolve().parents[3] / "data" / "exports"
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = output_dir / f"rostercopiilot_schedule_{ts}.xlsx"
    wb = build_workbook(result, include_audit=include_audit)
    wb.save(path)
    return path

